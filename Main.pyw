#BIBLIOTECAS DE INTERFACE GRÁFICA
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from threading import *
import pandas as pd
from openpyxl import load_workbook
#IMPORT DAS MINHAS PRÓPRIAS FUNÇÕES
from Despertador import ler_Agenda 
from Despertador import agendar

#BIBLIOTECA PARA NÃO ABRIR TERMINAL DO PYTHON E FICAR MOSTRANDO TEXTO
import ctypes
import os

#FUNÇÃO QUE FINALIZA O KERNEL
ctypes.windll.kernel32.FreeConsole()

#IMPORT TIME PARA DAR ALGUNS SLEEPS
from time import sleep

#CRIANDO JANELA
janela = Tk()

#PARÂMETROS DE JANELA
#TÍTULO DA JANELA
janela.title("Controle de Grind")

#TAMANHO PADRONIZADO PARA APLICAÇÃO
width_jan = 635
heigth_jan = 330

#MÁXIMOS E MÍNIMOS DA FUNÇÃO
janela.minsize(width_jan,heigth_jan)
janela.maxsize(width_jan,heigth_jan)

#CENTRALIZANDO TELA NO MONITOR
#COMO UTILIZO ESSE TRECHO EM MUITOS CÓDIGOS VOU CRIAR UMA FUNÇÃO PARA UTILIZAR POSTERIORMENTE
def centraliza_TK(jan,jan_width,jan_height):
    #ENCONTRO DIMENÇÕES DO MONITOR   
    mon_width = janela.winfo_screenwidth()
    mon_height = janela.winfo_screenheight()
    #DETERMINO POS X ATRAVÉS 
    #DO PONTO MEDIO MONITOR E PONTO MÉDIO DA MINHA JANA
    pos_x = (mon_width/2)-(jan_width/2)
    #DETERMINO POS y ATRAVÉS 
    #DO PONTO MEDIO MONITOR E PONTO MÉDIO DA MINHA JANA
    pos_y = (mon_height/2)-(jan_height/2)
    jan.geometry("%dx%d+%d+%d"%(jan_width,jan_height,pos_x,pos_y))
centraliza_TK(janela,width_jan,heigth_jan)

#FUNÇÕOES DOS BOTÕES:
#FUNÇÃO #1 PARA SELECIONAR A AGENDA DE TORNEIOS
def escolherAgenda(): 
    global agenda

    #FUNÇÃO TK INTER PARA ABRIR A PAGINA DE SELECIONAR ARQUIVOS DO TIPO EXCEL CSV OU ANY FILE
    try:
        agenda = filedialog.askopenfilename(filetypes=(("Arquivos Excel", "*.xlsx"), ("Any file", "*")))
        carregaSheetsName(agenda)
        return agenda
    except:
        entry_SheetName["state"]=DISABLED
        agenda = None
        for i in treeTorneios.get_children():
            treeTorneios.delete(i)
        janela.update()
        messagebox.showerror(title="Modelo de Pasta de Trabalho Incompatível ou Vazia", message="Selecione uma Pasta de Trabalho do Tipo .XLSX para carregar sua agenda de grind")
       
def carregaSheetsName(pasta_agenda):
    valores_SheetName = load_workbook(pasta_agenda, read_only=True).sheetnames
    entry_SheetName.configure(values=valores_SheetName, state=NORMAL)
    entry_SheetName.current(0)

def carregarAgenda():
    global silver_plan 

        #LIMPO A TABELA PARA REMOVER TORNEIOS ANTIGOS QUE JÁ FORAM CARREGADOS, OU ATÉ MESMO SE O USUÁRIO CARREGOU UM ARQUIVO ERRADO ANTERIORMENTE
    for i in treeTorneios.get_children():
        treeTorneios.delete(i)

    #ATUALIZO A JANELA PARA INICIAR O CARREGAMENTO DOS NOVOS DADOS
    janela.update()
       
    try:
        #FAÇO A LEITURA DA AGENDA DE TORNEIOS DO DIA
        silver_plan = ler_Agenda(agenda,entry_SheetName.get())
        #remover as colunas de hora de inicio e estágio de torneio
        silver_plan = silver_plan.drop(["Hora Inicio","Estágio Torneio"],axis="columns")      
        
        #PREENCHO A PLANILHA
        for index, row in  silver_plan.iterrows():
            treeTorneios.insert("", 'end', text=index, values=list(row))
            
        #HABILITAR BOTÃO DE INICIAR GRIND
        botaoIniciarDesp['state']=NORMAL 

    except:
        botaoIniciarDesp['state']=DISABLED 
        messagebox.showerror(title="Planilha Selecionada Incompatível",message="A planilha selecionada está fora dos padrões desejáveis, revise a planilha e deixe a nos padrões de execução do sistema. \nRevise os CABEÇALHOS, PLANILHA E PASTA DE TRABALHO selecionados.\nCaso não encontre erros comunique ao desenvolvedor")
        pass

def threading(plan,treeV):     
    t1=Thread(target=lambda: agendar(plan,treeV)) 
    t1.start() 

def finalizar_processo():
    janela.destroy()
    os._exit(0)


#FRAME 1 - CARREGAR AGENDA
text_Export = Label(janela, text="Escolha a pasta de trabalho que contém a grade atualizada: ")
text_Export.grid(column=0,row=0, sticky='w')

botaoEscolherAgenda = Button(janela, text="Buscar", command=escolherAgenda)
botaoEscolherAgenda.grid(column=0,row=0, sticky='e')

text_SheetName = Label(janela, text="Escolha a planilha que contém a grade do dia: ")
text_SheetName.grid(column=0,row=1,sticky='w')

entry_SheetName = ttk.Combobox(janela, state=DISABLED)
entry_SheetName.grid(column=0,row=2,sticky='w')


botaoEscolherAgenda = Button(janela, text="Carregar", command=carregarAgenda)
botaoEscolherAgenda.grid(column=0,row=2,sticky='e')

####

#PARÂMETROS DA TABELA
#CRIANDO TABELA E ESTABELECENDO QUANTIDADE DE COLUNAS
cols = ["columnn1", "columnn2", "columnn3", "columnn4", "columnn5", "columnn6"]
treeTorneios = ttk.Treeview(janela, selectmode="browse",show="headings", columns=cols, height=10)

#PARÂMETROS DA PRIMEIRA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn1",width=100, minwidth=50, stretch=NO)
treeTorneios.heading("#1", text='Hora de Registro')

#PARÂMETROS DA SEGUNDA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn2",width=50, minwidth=50, stretch=NO)
treeTorneios.heading("#2", text='Buy In')

#PARÂMETROS DA TERCEIRA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn3",width=70, minwidth=50, stretch=NO)
treeTorneios.heading("#3", text='GTD')

#PARÂMETROS DA QUARTA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn4",width=250, minwidth=50, stretch=NO)
treeTorneios.heading("#4", text='Descrição')

#PARÂMETROS DA QUINTA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn5",width=80, minwidth=50, stretch=NO)
treeTorneios.heading("#5", text='Site')

#PARÂMETROS DA SEXTA COLUNA, ESTABELECENDO ÍNDICE E CABEÇALHO
treeTorneios.column("columnn6",width=80, minwidth=50, stretch=NO)
treeTorneios.heading("#6", text='Prioridade')

#POSIÇÃO DA TABELA DENTRO DA JANELA
treeTorneios.grid(column=0,row=3,sticky='n')


#INICIANDO DESPERTADOR DE TORNEIOS:
botaoIniciarDesp = Button(janela, text="Iniciar Grind", command=lambda: threading(silver_plan,treeTorneios),state=DISABLED)
botaoIniciarDesp.grid(column=0,row=4,sticky='n')

#ENCERRANDO DESPERTADOR DE TORNEIOS:
botaoFinDesp = Button(janela, text="Finalizar Grind", command=finalizar_processo)
botaoFinDesp.grid(column=0,row=4,sticky='e')


#CARREGANDO DADOS DA TABELA
#arquivo = "C:\\Users\\jvBra\\OneDrive\\Documents\\2023\\Python Scripts\\SilverPlan\\Data\\Torneios.xlsx"
#sheet = "SilverPlan_2023_Domingo"

#INICIANDO JANELA PRINCIPAL
janela.mainloop()
